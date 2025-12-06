from __future__ import annotations

import os
from typing import Iterable, List, Set, Tuple

from openpyxl import load_workbook


def normalize_href(u: str) -> str:
    """
    중복판정 안정화:
    - 공백 제거
    - #fragment 제거
    - ?query 제거
    """
    if not u:
        return ""
    u = str(u).strip()
    u = u.split("#", 1)[0]
    u = u.split("?", 1)[0]
    return u


def load_hrefs_from_xlsx(path: str, href_col: int = 1) -> Set[str]:
    """
    엑셀 1개에서 href_col(기본 A열=1) 값을 모두 읽어서 set으로 반환.
    - 하이퍼링크 셀이면 hyperlink.target 우선
    - 헤더가 'href'면 자동 스킵
    """
    hrefs: Set[str] = set()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    try:
        for row in ws.iter_rows(min_row=1, min_col=href_col, max_col=href_col):
            cell = row[0]
            s = ""

            # 하이퍼링크 target 우선
            try:
                if getattr(cell, "hyperlink", None) and cell.hyperlink.target:
                    s = str(cell.hyperlink.target).strip()
            except Exception:
                pass

            # 없으면 셀 값 사용
            if not s and cell.value is not None:
                s = str(cell.value).strip()

            if not s:
                continue
            if s.lower() == "href":  # 헤더 스킵
                continue

            s = normalize_href(s)
            if s:
                hrefs.add(s)

    finally:
        wb.close()

    return hrefs


def load_hrefs_from_many(
    paths: Iterable[str],
    max_files: int = 100,
    href_col: int = 1,
) -> Tuple[Set[str], List[str], List[Tuple[str, str]]]:
    """
    여러 엑셀에서 href를 합쳐 로드.
    반환: (href_set, used_paths, errors[(path, reason)])
    """
    used: List[str] = []
    errors: List[Tuple[str, str]] = []
    merged: Set[str] = set()

    # 중복 제거 + 존재 체크
    uniq: List[str] = []
    seen = set()
    for p in paths or []:
        p = str(p)
        if p in seen:
            continue
        seen.add(p)
        uniq.append(p)

    uniq = uniq[:max_files]

    for p in uniq:
        if not os.path.exists(p):
            errors.append((p, "파일이 존재하지 않음"))
            continue
        if not p.lower().endswith(".xlsx"):
            errors.append((p, "xlsx 파일이 아님"))
            continue

        try:
            merged.update(load_hrefs_from_xlsx(p, href_col=href_col))
            used.append(p)
        except Exception as e:
            errors.append((p, f"로드 실패: {e}"))

    return merged, used, errors
